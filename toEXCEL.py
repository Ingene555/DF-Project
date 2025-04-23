""" 
This library will be used for conversion into excel
A class per file type will be defined
"""

import mainFunctions as mf
import os
from datetime import datetime as dt
import csv
from openpyxl import Workbook, load_workbook
import chardet
import openpyxl
import pandas as pd

TEMPCACHES = "tempcaches"


def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
        return result['encoding']


class ExcelToExcel:

    def __init__(self, file):
        self.file = file if type(file) == list else [file]
        self.stop = False
        self.progressShow = None

    def Convert(self, pathF, pathT):
        pathT
        _ = list()
        for ww in range(len(pathF)):
            _.append(os.path.join(TEMPCACHES, f"tc-excelel-{ww}-{dt.now()}.xlsx".replace(":", "")))
        while(len(pathF) > len(_)):
            _.append(os.path.join(TEMPCACHES, f"tc-excelel-{dt.now()}.xlsx".replace(":", "")))
        __ = list()
        for xx, ww in enumerate(pathF):
            if not self.stop:
                if self.progressShow:
                    self.progressShow(int(xx * 100 / len(pathF)))
                try:
                    ext = os.path.splitext(ww)[-1].lower()
                    print(ext)
                    if ext == ".csv":
                        wb = Workbook()
                        ws = wb.active
                        encoding = detect_encoding(ww)
                        with open(ww, mode='r', encoding=encoding) as f:
                            reader = csv.reader(f)
                            for row in reader:
                                ws.append(row)
                        wb.save(_[xx])
                    elif ext == ".xlsx":
                        wb = load_workbook(ww)
                        wb.save(_[xx])
                    elif ext == ".xlsb":
                        with pd.ExcelFile(ww, engine="pyxlsb") as xlsb:
                            with pd.ExcelWriter(_[xx], engine="openpyxl") as writer:
                                for yy, sheet_name in enumerate(xlsb.sheet_names):
                                    df = xlsb.parse(sheet_name)
                                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                                    if self.progressShow:
                                        self.progressShow(int(yy * 100 / len(xlsb.sheet_names)))
                    else:
                        excel_data = pd.ExcelFile(ww)
                        with pd.ExcelWriter(_[xx], engine='openpyxl') as writer:
                            for yy, sheet_name in enumerate(excel_data.sheet_names):
                                df = excel_data.parse(sheet_name)
                                df.to_excel(writer, index=False, sheet_name=sheet_name)
                                if self.progressShow:
                                    self.progressShow(int(yy * 100 / len(excel_data.sheet_names)))
                    __.append(mf.setReturn(True, [type(_[xx])], [_[xx]]))
                except Exception as e:
                    __.append(mf.setReturn(False, [type(e)], [e]))
            else:
                for i in __:
                    try:
                        os.remove(mf.Return(i).getValues(0))
                    except:pass
                break
        if self.progressShow:
            self.progressShow(100)
        return __

    def toEXCEL(self, path):
        temp = self.Convert(self.file, [])
        path = path if type(path) == list else [path]
        _ = list()
        for xx, i in enumerate(temp):
            if not self.stop:
                ww = mf.Return(i).getValues(0)
                if mf.Return(i).getSucced():
                    ext = os.path.splitext(path[xx])[-1].lower()
                    try:
                        if ext == ".csv":
                            workbook = openpyxl.load_workbook(ww)
                            sheet = workbook.active
                            with open(path[xx], mode='w', newline='', encoding='utf-8') as f:
                                writer = csv.writer(f)
                                for row in sheet.iter_rows(values_only=True):
                                    writer.writerow(row)
                        else:
                            wb = load_workbook(ww)
                            wb.save(path[xx])
                        _.append(mf.setReturn(True, [type(path[xx])], [path[xx]]))
                    except Exception as e:
                        _.append(mf.setReturn(False, [type(e)], [e]))
                else:_.append(i)
            else:
                for i in _:
                    try:
                        os.remove(mf.Return(i).getValues(0))
                    except:pass
                _=[]
                break
        if self.progressShow:
            self.progressShow(100, True, self.stop, _)
        return _

    def kill(self):
        self.stop = True

